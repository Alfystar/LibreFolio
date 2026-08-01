"""Shared IO and number/date parsing helpers for Italian-broker BRIM plugins.

This module is intentionally **not** a provider: its filename starts with ``_``,
so ``BRIMProviderRegistry.auto_discover`` skips it (see
``backend/app/services/provider_registry.py``). It is imported directly by the
plugins that need it (Intesa Sanpaolo, Crédit Agricole), both of which
must read the *same* logical export from either **CSV** or **XLSX**.

Responsibilities:
- Read a tabular file (CSV or XLSX) into a uniform ``list[list[cell]]``.
  * CSV cells are always ``str``.
  * XLSX cells keep their native openpyxl type (``str`` / ``int`` / ``float`` /
    ``datetime`` / ``None``) because different exports store numbers either as
    native floats (Intesa *lista*) or as Italian-formatted text (Intesa
    *patrimonio*, Crédit Agricole).
- Locate a header row dynamically and map header labels to column indexes, so a
  plugin never hardcodes a fixed column offset (the same logical export has the
  data starting at column A in CSV but column C in the XLSX variant).
- Parse Italian-formatted numbers (``1.234,56`` -> ``1234.56``) and plain/Anglo
  numbers (``1867.178`` -> ``1867.178``) and multi-format dates.

None of these helpers touch the FX subsystem or recompute any figure; they only
transcribe and normalise what the broker reported (BRIM verbatim rule).
"""

from __future__ import annotations

import csv
import warnings
from dataclasses import dataclass
from datetime import date as date_type
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

import openpyxl

XLSX_EXTENSIONS = {".xlsx", ".xlsm"}
CSV_EXTENSIONS = {".csv", ".txt"}

# Encodings tried in order when reading CSV/text (Windows exports often ship BOM).
_TEXT_ENCODINGS = ("utf-8-sig", "utf-8", "latin-1", "cp1252")

# Symbols stripped before numeric parsing.
_CURRENCY_SYMBOLS = ("€", "$", "£", "%", "'")


def is_xlsx(file_path: Path) -> bool:
    """True when the file extension denotes an Excel workbook we can read."""
    return file_path.suffix.lower() in XLSX_EXTENSIONS


# ---------------------------------------------------------------------------
# File reading (CSV or XLSX -> list[list[cell]])
# ---------------------------------------------------------------------------


def read_rows(file_path: Path, *, delimiter: Optional[str] = None) -> List[List[Any]]:
    """Read a CSV or XLSX file into a list of rows (each a list of raw cells).

    For CSV the delimiter is auto-detected when not supplied. For XLSX the first
    worksheet that contains any data is read with ``data_only=True`` (so formula
    results, not formulas, are returned).
    """
    if is_xlsx(file_path):
        return _read_xlsx_rows(file_path)
    return _read_csv_rows(file_path, delimiter=delimiter)


def _read_csv_rows(file_path: Path, *, delimiter: Optional[str]) -> List[List[Any]]:
    if delimiter is None:
        delimiter = detect_delimiter(file_path)
    last_err: Optional[Exception] = None
    for enc in _TEXT_ENCODINGS:
        try:
            with open(file_path, encoding=enc, newline="") as f:
                return [list(row) for row in csv.reader(f, delimiter=delimiter)]
        except UnicodeDecodeError as exc:  # try next encoding
            last_err = exc
            continue
    if last_err is not None:
        raise last_err
    return []


def _read_xlsx_rows(file_path: Path) -> List[List[Any]]:
    with warnings.catch_warnings():
        # openpyxl emits harmless UserWarnings for some broker exports
        # (missing default style, unsupported print areas).
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=False)
        try:
            worksheet = _pick_worksheet(wb)
            if worksheet is None:
                return []
            return [list(row) for row in worksheet.iter_rows(values_only=True)]
        finally:
            wb.close()


def _pick_worksheet(workbook: Any) -> Any:
    """Return the first worksheet that holds at least one non-empty cell."""
    for ws in workbook.worksheets:
        for row in ws.iter_rows(values_only=True):
            if any(cell not in (None, "") for cell in row):
                return ws
    return workbook.active


def detect_delimiter(file_path: Path, lines_to_read: int = 15) -> str:
    """Sniff the CSV delimiter (``,`` / ``;`` / tab) with a safe fallback."""
    sample = ""
    for enc in _TEXT_ENCODINGS:
        try:
            with open(file_path, encoding=enc) as f:
                for i, line in enumerate(f):
                    if i >= lines_to_read:
                        break
                    sample += line
            break
        except UnicodeDecodeError:
            sample = ""
            continue
    try:
        return csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t"]).delimiter
    except Exception:
        return ";" if sample.count(";") > sample.count(",") else ","


# ---------------------------------------------------------------------------
# Header location and column mapping
# ---------------------------------------------------------------------------


def cell_str(value: Any) -> str:
    """Stringify a raw cell for label comparison (``None`` -> empty string)."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def find_header_row(
    rows: Sequence[Sequence[Any]],
    required_labels: Sequence[str],
    *,
    max_scan: int = 60,
) -> Optional[int]:
    """Index of the first row containing *all* ``required_labels``.

    Matching is case-insensitive substring against each stringified cell, so a
    header cell ``"Categoria "`` still matches the label ``"categoria"``.
    """
    req = [r.strip().lower() for r in required_labels]
    for i, row in enumerate(rows):
        if i >= max_scan:
            break
        cells = [cell_str(c).lower() for c in row]
        if all(any(r in c for c in cells) for r in req):
            return i
    return None


def build_col_index(header_row: Sequence[Any], mapping: Dict[str, Sequence[str]]) -> Dict[str, int]:
    """Map logical keys to column indexes using the real header labels.

    ``mapping`` is ``{logical_key: [candidate labels]}``. The first exact
    (case-insensitive) label match wins, and the first matching column is kept
    (important when a label such as ``Divisa`` appears more than once and the
    first occurrence is the one we want).
    """
    normalized = [cell_str(c).lower() for c in header_row]
    result: Dict[str, int] = {}
    for key, candidates in mapping.items():
        for cand in candidates:
            target = cand.strip().lower()
            for col, cell in enumerate(normalized):
                if cell == target:
                    result[key] = col
                    break
            if key in result:
                break
    return result


def row_get(row: Sequence[Any], col_index: Dict[str, int], key: str) -> Any:
    """Safely fetch ``row[col_index[key]]``; ``None`` when missing/out of range."""
    idx = col_index.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def is_blank_row(row: Sequence[Any]) -> bool:
    """True when every cell in the row is empty/blank."""
    return all(cell_str(c) == "" for c in row)


# ---------------------------------------------------------------------------
# Number and date parsing
# ---------------------------------------------------------------------------


def _coerce_number(value: Any, *, decimal_sep: str, thousands_sep: str) -> Optional[Decimal]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    s = str(value).strip()
    for sym in _CURRENCY_SYMBOLS:
        s = s.replace(sym, "")
    s = s.replace("\xa0", "").replace(" ", "").strip()
    if s in ("", "-", "–", "—"):
        return None
    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1]
    if thousands_sep:
        s = s.replace(thousands_sep, "")
    if decimal_sep and decimal_sep != ".":
        s = s.replace(decimal_sep, ".")
    try:
        result = Decimal(s)
    except InvalidOperation:
        return None
    return -result if negative else result


def to_decimal_it(value: Any) -> Optional[Decimal]:
    """Parse an Italian-formatted number: ``.`` = thousands, ``,`` = decimal.

    Accepts native ``int``/``float``/``Decimal`` unchanged and tolerates trailing
    currency symbols/spaces (``"50.683,13"`` -> ``50683.13``,
    ``"107,7662 €"`` -> ``107.7662``, ``"9.991"`` -> ``9991``).
    """
    return _coerce_number(value, decimal_sep=",", thousands_sep=".")


def to_decimal_plain(value: Any) -> Optional[Decimal]:
    """Parse a plain/Anglo number: ``.`` = decimal, optional ``,`` thousands.

    Used for columns that use dotted decimals even in an otherwise Italian file
    (e.g. Crédit Agricole fund ``Quantità`` such as ``"1867.178"``; integer
    bond nominals such as ``"50000"`` parse unchanged).
    """
    return _coerce_number(value, decimal_sep=".", thousands_sep=",")


_DATE_FORMATS = ("%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%d.%m.%y")


def to_date(value: Any, formats: Sequence[str] = _DATE_FORMATS) -> Optional[date_type]:
    """Parse a date from a native datetime/date or a string in common IT formats."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date_type):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# Bond redemption mechanics are shared across Italian-broker plugins: a bond is
# quoted per 100 of nominal and redeemed at par, whereas a fund is priced per unit.
# ``looks_like_bond`` is the single heuristic used to decide whether the maturity /
# redemption "at par + surplus-as-INTEREST" model applies (see ``model_bond_maturity``).
_BOND_KEYWORDS = ("BTP", "BOT", "CCT", "CTZ", "BUND", "OAT", "OBBLIG", "CCTEU", "BOND")


def looks_like_bond(name: Optional[str], isin: Optional[str] = None) -> bool:
    """Heuristic: does this security follow bond price conventions (quoted per 100)?

    Matches common Italian / European government-bond mnemonics in the security name
    (``BTP``, ``BOT``, ``CCT``, ``BUND`` …). ``isin`` is accepted for future refinement
    but is not required — the name keywords are sufficient for the current exports.
    """
    if name:
        upper = name.upper()
        if any(kw in upper for kw in _BOND_KEYWORDS):
            return True
    return False


@dataclass(frozen=True)
class BondMaturityModel:
    """Result of modelling a bond redeemed at maturity (``TITOLI SCADUTI`` etc.).

    A bond is always redeemed **at par** (100). Any amount the bank credits *above*
    par is a *premio fedeltà* / inflation revaluation (e.g. a BTP Italia ``FOI``
    indexation) which is reddito di capitale, and must be booked as INTEREST rather
    than folded into the sale price. Splitting keeps the SELL leg exactly at par so
    the position closes cleanly and the realised gain reflects price-vs-cost only.
    """

    nominal: Decimal
    """Face value to redeem — the magnitude of the SELL quantity (always ``>= 0``)."""
    principal_cash: Decimal
    """Cash attributable to the par redemption (``nominal`` valued at ``par``)."""
    surplus_cash: Decimal
    """Amount credited above par (``>= 0``) — to be booked as INTEREST income."""
    par_price: Decimal
    """Unit price of the SELL leg (par, e.g. 100 for a bond quoted per 100)."""
    source: str
    """``"position"`` when ``nominal`` came from the held position, ``"derived"``
    when it had to be inferred from ``ctv / price`` (the export reported quantity 0
    and no prior position was found — the caller should then attach a *verify*
    field-todo, because the SELL is the only carrier of the nominal)."""


def model_bond_maturity(
    *,
    ctv: Decimal,
    price: Optional[Decimal],
    held_qty: Optional[Decimal] = None,
    par: Decimal = Decimal(100),
) -> BondMaturityModel:
    """Model a bond redemption at maturity, splitting par principal from surplus.

    ``ctv`` is the total cash the bank credited for the redemption; ``price`` is the
    reported redemption price per ``par`` of nominal (it may embed the premio /
    revaluation, e.g. ``100.40``); ``held_qty`` is the net nominal currently held for
    the security (from prior BUY / succession / seed rows).

    Working assumption (documented per-plugin): a redeemed security is a bond quoted
    **at par (100)**, so anything credited above par is income. Rules:

    * ``held_qty > 0`` -> ``nominal = held_qty`` (exact position close),
      ``source = "position"``. This is preferred: the true nominal lives in the
      position, never in the maturity row (whose quantity is 0).
    * otherwise -> ``nominal = ctv / price * par`` (best-effort derivation),
      ``source = "derived"``. The caller should flag such a row for the user to
      verify, since the SELL is then the only carrier of the nominal.
    * ``principal_cash = nominal * par / 100`` (the par redemption value).
    * ``surplus_cash = ctv - principal_cash`` (``>= 0``). A price below par (rare)
      yields no surplus: everything stays principal so no negative income is invented.
    """
    ctv = abs(ctv)
    if held_qty is not None and held_qty > 0:
        nominal = held_qty
        source = "position"
    elif price:
        nominal = ctv / price * par
        source = "derived"
    else:
        # No position and no usable price: cannot model a split. Return a zero
        # nominal so the caller falls back to its own (verbatim) handling.
        return BondMaturityModel(nominal=Decimal(0), principal_cash=ctv, surplus_cash=Decimal(0), par_price=par, source="derived")

    nominal = nominal.quantize(Decimal("0.001"))
    principal = (nominal * par / Decimal(100)).quantize(Decimal("0.01"))
    surplus = ctv - principal
    if surplus < 0:
        principal = ctv
        surplus = Decimal(0)
    return BondMaturityModel(nominal=nominal, principal_cash=principal, surplus_cash=surplus, par_price=par, source=source)


# ---------------------------------------------------------------------------
# Maturity / redemption detection (drives BRIMAssetNotice on extracted assets)
# ---------------------------------------------------------------------------

MATURITY_NOTICE_KIND = "maturity_suspected"
"""``BRIMAssetNotice.kind`` used when a maturity/redemption movement is detected.

Kept here (shared) so every Italian-broker plugin emits the *same* category key and the
frontend can group and label these notices consistently."""

_MATURITY_KEYWORDS = ("scadut", "scaden", "rimbors", "estinzion", "redemption", "matured", "maturity")
"""Substrings (matched case-insensitively) that flag a redeemed / matured security.

Grounded on the real Italian exports: ``TITOLI SCADUTI`` (-> ``scadut``) and
``FONDI: RIMBORSO`` (-> ``rimbors``). The English terms are for robustness/future exports."""


def looks_like_maturity(text: Optional[str]) -> bool:
    """Heuristic: does this free-text (a causale / operazione / description) imply a redemption?"""
    if not text:
        return False
    lowered = text.lower()
    return any(kw in lowered for kw in _MATURITY_KEYWORDS)


def detect_maturity_hits(transactions: Sequence[Any]) -> Dict[int, List[int]]:
    """Scan parsed transactions and group, per asset, the indexes whose description looks
    like a maturity/redemption.

    Returns ``{asset_id: [tx_index, ...]}`` (only assets with at least one hit). Decoupled
    from the schema types on purpose (uses ``getattr``) so this IO module keeps no schema
    imports. Callers turn each entry into a ``BRIMAssetNotice`` with a provider-localized reason.
    """
    hits: Dict[int, List[int]] = {}
    for idx, tx in enumerate(transactions):
        asset_id = getattr(tx, "asset_id", None)
        if asset_id is None:
            continue
        if looks_like_maturity(getattr(tx, "description", None)):
            hits.setdefault(asset_id, []).append(idx)
    return hits
